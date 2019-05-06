import openpyxl

def read_excel(filename, sheetname, row_numb, col_numb, with_head):
  objects = []
  wb = openpyxl.load_workbook(filename=filename)
  ws = wb[sheetname]
  attrs = []
  row_st = 1
  if with_head:
    for i in range(1, col_numb + 1):
      attrs.append(str(ws.cell(row=1, column=i).value))
    row_st = 2

  for i in range(row_st, row_numb + 1):
    if with_head:
      row = {}
    else:
      row = []
    for j in range(1, col_numb + 1):
      data = str(ws.cell(row=i, column=j).value)
      if with_head:
        row[attrs[j - 1]] = data
      else:
        row.append(data)
    objects.append(row)
  return objects

def write_excel(filename, datas, sheetname):
  row_numb = len(datas)
  col_numb = 0 if row_numb == 0 else len(datas[0])
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = sheetname
  for i in range(1, row_numb + 1):
    for j in range(1, col_numb + 1):
      ws.cell(row=i, column=j).value = datas[i - 1][j - 1]
  wb.save(filename)

def main():
  filename = 'code/analysis.out'
  f = open(filename, 'r')
  lines = list(map(lambda x: x.strip(), f.readlines()))
  f.close()
  datas = []
  for i in range(0, len(lines), 2):
    graph = lines[i].split(' ')[1].split('.')[0]
    kvs = list(map(lambda x: x.split(':'), lines[i + 1].split(' ')))
    datas.append([graph, kvs[3][1]])

  output_filename = 'analysis.xlsx'
  write_excel(output_filename, datas, 'LayerCount')

if __name__ == '__main__':
  main()